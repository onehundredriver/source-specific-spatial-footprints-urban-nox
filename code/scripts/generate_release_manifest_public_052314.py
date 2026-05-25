# -*- coding: utf-8 -*-
"""
Generate release_manifest_public.xlsx for the revised Nature Cities release package.

Current release structure:
- Main figures: Fig. 1鈥?
- Extended Data figures: Extended Data Fig. 1鈥?
- Supplementary figures: Supplementary Fig. 1鈥?
"""

from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OUT_XLSX = ROOT / "release_manifest_public.xlsx"

HEADERS = [
    "Item_ID",
    "Item_Type",
    "Relative_Path",
    "Required",
    "Access",
    "Description",
    "Relevant_Manuscript_Section",
    "Status",
    "Size_MB",
    "Notes",
]

rows = []


def size_mb(rel_path):
    p = ROOT / rel_path
    if p.exists() and p.is_file():
        return round(p.stat().st_size / 1024 / 1024, 4)
    if p.exists() and p.is_dir():
        total = sum(x.stat().st_size for x in p.rglob("*") if x.is_file())
        return round(total / 1024 / 1024, 4)
    return ""


def exists_status(rel_path):
    return "Included" if (ROOT / rel_path).exists() else "Missing"


def add(item_id, item_type, rel_path, required, access, desc, section, status=None, notes=""):
    rows.append([
        item_id,
        item_type,
        rel_path,
        required,
        access,
        desc,
        section,
        status or exists_status(rel_path),
        size_mb(rel_path),
        notes,
    ])


# ---------------------------------------------------------------------
# Documentation and top-level code
# ---------------------------------------------------------------------
add(
    "README",
    "Documentation",
    "README.md",
    "Yes",
    "Public",
    "Repository overview, included materials and usage instructions.",
    "Data availability; Code availability",
)

add(
    "REPRODUCIBILITY_GUIDE",
    "Documentation",
    "REPRODUCIBILITY_GUIDE.md",
    "Yes",
    "Public",
    "Guide for reproducing released figures and model workflows.",
    "Code availability",
)

add(
    "RUN_ALL_FIGURES",
    "Code",
    "code/run_all_figure_reproductions.py",
    "Yes",
    "Public",
    "Convenience script to reproduce all main, Extended Data and supplementary figures from released source-data files.",
    "Code availability",
)

add(
    "INTEGRITY_CHECK",
    "Code",
    "code/check_release_package_integrity.py",
    "Recommended",
    "Public",
    "Integrity checker for the revised release package.",
    "Code availability",
)

add(
    "RELEASE_INTEGRITY_REPORT",
    "Documentation",
    "docs/release_integrity_check.md",
    "Recommended",
    "Public",
    "Latest release integrity check report.",
    "Code availability",
)


# ---------------------------------------------------------------------
# Main figures
# ---------------------------------------------------------------------
add(
    "FIG1_SOURCE_DIR",
    "Figure source data",
    "data/source_data/main/Fig1",
    "Yes",
    "Public",
    "Compact source-data directory for Fig. 1, including monitoring sites, road-network traffic-flow proxy and schematic parameters.",
    "Fig. 1; Data availability",
)

add(
    "CODE_FIG1",
    "Code",
    "code/scripts/reproduce_fig1_from_source_data.py",
    "Yes",
    "Public",
    "Standalone script to reproduce Fig. 1 from released source data.",
    "Fig. 1; Code availability",
)

for ext in ["png", "pdf"]:
    add(
        f"FIG1_REPRO_{ext.upper()}",
        "Main Figure",
        f"figures/main/Fig1_reproduced_from_source_data.{ext}",
        "Recommended",
        "Public",
        f"{ext.upper()} version of Fig. 1 reproduced from released source data.",
        "Fig. 1; Code availability",
    )

for i, desc in [
    (2, "Plot-ready source data for source-specific spatial footprint diagnostics."),
    (3, "Plot-ready source data for Buffer versus Aggregated benchmark comparison."),
    (4, "Plot-ready source data for micro-environmental transferability typologies."),
]:
    fig = f"Fig{i}"
    add(
        f"FIG{i}_SOURCE",
        "Figure source data",
        f"data/source_data/main/{fig}_source_data.csv",
        "Yes",
        "Public",
        desc,
        f"Fig. {i}; Data availability",
    )
    if i == 4:
        add(
            "FIG4_TRAFFIC_NETWORK",
            "Figure source data",
            "data/source_data/main/Fig4_traffic_network.csv",
            "Yes",
            "Public",
            "Road-network traffic-flow proxy source data used in Fig. 4 map panel.",
            "Fig. 4; Data availability",
        )
    add(
        f"CODE_FIG{i}",
        "Code",
        f"code/scripts/reproduce_fig{i}_from_source_data.py",
        "Yes",
        "Public",
        f"Standalone script to reproduce Fig. {i} from released source data.",
        f"Fig. {i}; Code availability",
    )
    for ext in ["png", "pdf"]:
        add(
            f"FIG{i}_REPRO_{ext.upper()}",
            "Main Figure",
            f"figures/main/{fig}_reproduced_from_source_data.{ext}",
            "Recommended",
            "Public",
            f"{ext.upper()} version of Fig. {i} reproduced from released source data.",
            f"Fig. {i}; Code availability",
        )

add(
    "FIG5_SOURCE_DIR",
    "Figure source data",
    "data/source_data/main/Fig5",
    "Yes",
    "Public",
    "Compact sampled plot-ready source-data directory for full main Fig. 5.",
    "Fig. 5; Data availability",
    notes="Full SHAP and validation matrices are not bundled in GitHub; sampled plot-ready points are released for figure reproduction.",
)

add(
    "CODE_FIG5",
    "Code",
    "code/scripts/reproduce_fig5_from_source_data.py",
    "Yes",
    "Public",
    "Standalone script to reproduce full main Fig. 5 from compact sampled source data.",
    "Fig. 5; Code availability",
)

for ext in ["png", "pdf"]:
    add(
        f"FIG5_REPRO_{ext.upper()}",
        "Main Figure",
        f"figures/main/Fig5_reproduced_from_source_data.{ext}",
        "Recommended",
        "Public",
        f"{ext.upper()} version of Fig. 5 reproduced from released source data.",
        "Fig. 5; Code availability",
    )

# Optional Fig. 5 chord component provenance
add(
    "FIG5_CHORD_SOURCE",
    "Figure source data",
    "data/source_data/main/Fig5_chord_source_data.csv",
    "Optional",
    "Public",
    "Component-level source data for the Fig. 5 chord-network panels.",
    "Fig. 5; Data availability",
    notes="Provided as component provenance; full Fig. 5 source data are stored under data/source_data/main/Fig5/.",
)

add(
    "CODE_FIG5_CHORD_COMPONENT",
    "Code",
    "code/scripts/reproduce_fig5_chord_component_from_source_data.py",
    "Optional",
    "Public",
    "Component script for reproducing Fig. 5 chord-network panels.",
    "Fig. 5; Code availability",
    notes="Component provenance only; main Fig. 5 should be reproduced with reproduce_fig5_from_source_data.py.",
)


# ---------------------------------------------------------------------
# Extended Data figures
# ---------------------------------------------------------------------
for i, desc in [
    (1, "Source-data workbook for robustness of source-specific footprint estimates."),
    (2, "Plot-ready source data for phase-space support of micro-environmental interpretation."),
    (3, "Plot-ready source data for the full marginal-contribution matrix."),
    (4, "Plot-ready source data for the full road-emission density 脳 wind-speed SHAP interaction matrix."),
    (5, "Plot-ready source data for the full road-emission density 脳 forest-density SHAP interaction matrix."),
    (6, "Plot-ready source data for the full road-emission density 脳 average-frontal-area SHAP interaction matrix."),
]:
    ext = "xlsx" if i == 1 else "csv"
    fig = f"Extended_Data_Fig{i}"
    add(
        f"EDF{i}_SOURCE",
        "Extended Data figure source data",
        f"data/source_data/extended_data/{fig}_source_data.{ext}",
        "Yes",
        "Public",
        desc,
        f"Extended Data Fig. {i}; Data availability",
        notes="" if i == 1 else "",
    )
    add(
        f"CODE_EDF{i}",
        "Code",
        f"code/scripts/reproduce_extended_data_fig{i}_from_source_data.py",
        "Yes",
        "Public",
        f"Standalone script to reproduce Extended Data Fig. {i}.",
        f"Extended Data Fig. {i}; Code availability",
    )
    for ext_out in ["png", "pdf"]:
        add(
            f"EDF{i}_REPRO_{ext_out.upper()}",
            "Extended Data Figure",
            f"figures/extended_data/{fig}_reproduced_from_source_data.{ext_out}",
            "Recommended",
            "Public",
            f"{ext_out.upper()} version of Extended Data Fig. {i} reproduced from released source data.",
            f"Extended Data Fig. {i}; Code availability",
        )


# ---------------------------------------------------------------------
# Supplementary figures 1鈥?
# ---------------------------------------------------------------------
for i in range(1, 9):
    fig = f"Supplementary_Fig{i}"
    csv_path = ROOT / f"data/source_data/supplementary/{fig}_source_data.csv"
    xlsx_path = ROOT / f"data/source_data/supplementary/{fig}_source_data.xlsx"
    if csv_path.exists():
        source = f"data/source_data/supplementary/{fig}_source_data.csv"
    elif xlsx_path.exists():
        source = f"data/source_data/supplementary/{fig}_source_data.xlsx"
    else:
        source = f"data/source_data/supplementary/{fig}_source_data.csv"

    add(
        f"SUPP_FIG{i}_SOURCE",
        "Supplementary figure source data",
        source,
        "Yes",
        "Public",
        f"Plot-ready source data for Supplementary Fig. {i}.",
        f"Supplementary Fig. {i}; Data availability",
    )
    add(
        f"CODE_SUPP_FIG{i}",
        "Code",
        f"code/scripts/reproduce_supplementary_fig{i}_from_source_data.py",
        "Yes",
        "Public",
        f"Standalone script to reproduce Supplementary Fig. {i}.",
        f"Supplementary Fig. {i}; Code availability",
    )
    for ext_out in ["png", "pdf"]:
        add(
            f"SUPP_FIG{i}_REPRO_{ext_out.upper()}",
            "Supplementary Figure",
            f"figures/supplementary/{fig}_reproduced_from_source_data.{ext_out}",
            "Recommended",
            "Public",
            f"{ext_out.upper()} version of Supplementary Fig. {i} reproduced from released source data.",
            f"Supplementary Fig. {i}; Code availability",
        )


# ---------------------------------------------------------------------
# Supplementary data and public data
# ---------------------------------------------------------------------
for i in [1, 2]:
    add(
        f"SUPP_DATA_{i}",
        "Supplementary Data",
        f"data/supplementary_data/Supplementary_Data_{i}.xlsx",
        "Yes",
        "Public",
        f"Supplementary Data {i} workbook.",
        f"Supplementary Data {i}; Data availability",
    )

for rel, desc in [
    ("data/public/national_control_station_metadata.csv", "Metadata for public national-control monitoring stations."),
    ("data/public/processed_national_control_NOx_hourly.csv", "Processed hourly NOx observations for public national-control monitoring stations."),
    ("data/public/NOx_QC_overall_summary.csv", "Overall NOx quality-control summary."),
    ("data/public/NOx_QC_summary_by_station.csv", "Station-level NOx quality-control summary."),
]:
    add(
        Path(rel).stem.upper(),
        "Public data",
        rel,
        "Yes",
        "Public",
        desc,
        "Data availability",
    )

for rel in [
    "data/splits/spatial_5fold_station_assignment.csv",
    "data/splits/temporal_5fold_window_assignment.csv",
    "data/splits/temporal_buffer_intervals.csv",
    "data/splits/temporal_split_metadata.csv",
    "data/splits/anomalous_station_exclusion_list.csv",
    "data/splits/final_cluster_station_labels.csv",
]:
    add(
        Path(rel).stem.upper(),
        "Split or station-label data",
        rel,
        "Yes",
        "Public",
        "Released split/station-label file supporting validation and typology analyses.",
        "Methods; Data availability",
    )


# ---------------------------------------------------------------------
# Model rerun workflows and configs
# ---------------------------------------------------------------------
for model, folder, prefix in [
    ("XGBoost", "xgboost", "xgboost"),
    ("Random Forest", "rf", "rf"),
    ("LightGBM", "lgbm", "lgbm"),
]:
    for track in ["spatial", "temporal"]:
        add(
            f"MODEL_{prefix.upper()}_{track.upper()}",
            "Model rerun workflow",
            f"code/pipelines/{folder}/run_{track}_cv.py",
            "Recommended",
            "Public",
            f"{model} {track} cross-validation rerun script.",
            "Code availability",
            notes="Large model-input matrices are hosted externally; outputs are generated by rerunning the workflow and are not bundled.",
        )
        for scale in ["1hour", "24hour", "168hour", "31day"]:
            add(
                f"CONFIG_{prefix.upper()}_{track.upper()}_{scale.upper()}",
                "Model config",
                f"code/config/{prefix}_{track}_{scale}.json",
                "Recommended",
                "Public",
                f"Configuration file for {model} {track} CV at {scale} aggregation.",
                "Code availability",
            )


# ---------------------------------------------------------------------
# Write workbook
# ---------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"
ws.append(HEADERS)

for row in rows:
    ws.append(row)

header_fill = PatternFill("solid", fgColor="D9EAF7")
header_font = Font(bold=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row in ws.iter_rows():
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

for col_idx, col in enumerate(ws.columns, start=1):
    max_len = 0
    for cell in col:
        val = "" if cell.value is None else str(cell.value)
        max_len = max(max_len, len(val))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 55)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# Metadata sheet
meta = wb.create_sheet("Metadata")
meta.append(["Field", "Value"])
meta.append(["Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
meta.append(["Repository", "Released GitHub package root"])
meta.append(["Figure structure", "Main Fig. 1鈥?; Extended Data Fig. 1鈥?; Supplementary Fig. 1鈥?"])
meta.append(["Notes", "Large model-input matrices are externally archived and are not intended to be committed to GitHub."])
for cell in meta[1]:
    cell.font = header_font
    cell.fill = header_fill

for col_idx, col in enumerate(meta.columns, start=1):
    max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    meta.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 70)

wb.save(OUT_XLSX)
print(f"[OK] Wrote {OUT_XLSX}")
print(f"[INFO] Manifest rows: {len(rows)}")
