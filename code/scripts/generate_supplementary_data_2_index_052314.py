# -*- coding: utf-8 -*-
"""
Generate Supplementary Data 2 for the revised Nature Cities release package.

Output:
    data/supplementary_data/Supplementary_Data_2.xlsx

Current figure structure:
- Main figures: Fig. 1鈥?
- Extended Data figures: Extended Data Fig. 1鈥?
- Supplementary figures: Supplementary Fig. 1鈥?
- Supplementary tables: 1鈥?
- Supplementary data: 1鈥?
- Model rerun workflows: 6
"""

from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "data" / "supplementary_data"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_XLSX = OUT_DIR / "Supplementary_Data_2.xlsx"


def exists(rel):
    if rel in ["Not applicable", "See notes", "Supplementary Information"]:
        return ""
    return "Yes" if (ROOT / rel).exists() else "No"


def size_mb(rel):
    p = ROOT / rel
    if p.exists() and p.is_file():
        return round(p.stat().st_size / 1024 / 1024, 4)
    return ""


rows = []


def add_row(item_id, item_type, label, description, source_data, script, output, inputs, level, notes):
    rows.append({
        "Item_ID": item_id,
        "Item_Type": item_type,
        "Current_Manuscript_Label": label,
        "Description": description,
        "Source_Data_Path": source_data,
        "Reproduction_Script_Path": script,
        "Expected_Output_Path": output,
        "Required_Inputs": inputs,
        "Reproducibility_Level": level,
        "Notes": notes,
        "Source_Data_Exists": exists(source_data),
        "Script_Exists": exists(script),
        "Output_Exists": exists(output),
        "Source_Data_Size_MB": size_mb(source_data),
        "Output_Size_MB": size_mb(output),
    })


# Main figures
add_row(
    "Fig1", "Main figure", "Fig. 1",
    "Study setting, fixed-radius-buffer contrast, wind-sensitive sector-grid construction and analytical workflow.",
    "data/source_data/main/Fig1/",
    "code/scripts/reproduce_fig1_from_source_data.py",
    "figures/main/Fig1_reproduced_from_source_data.png",
    "Compact station, road-network traffic-flow proxy and schematic parameter files under data/source_data/main/Fig1/.",
    "Reproducible from compact plot-ready source data",
    "New figure added in the revised Nature Cities manuscript."
)

for i, desc in [
    (2, "Source-specific spatial footprint diagnostics for vehicle, shipping and industrial emissions."),
    (3, "Buffer versus Aggregated benchmark comparison for spatial and temporal generalisation."),
    (4, "Micro-environmental transferability typologies and spatial context."),
]:
    add_row(
        f"Fig{i}", "Main figure", f"Fig. {i}", desc,
        f"data/source_data/main/Fig{i}_source_data.csv",
        f"code/scripts/reproduce_fig{i}_from_source_data.py",
        f"figures/main/Fig{i}_reproduced_from_source_data.png",
        "Compact plot-ready source data released in the repository.",
        "Reproducible from compact plot-ready source data",
        "Renumbered relative to the earlier manuscript figure set."
    )

add_row(
    "Fig5", "Main figure", "Fig. 5",
    "Wind-sensitive interactions between source attribution, meteorology, vegetation-related roughness and building morphology.",
    "data/source_data/main/Fig5/",
    "code/scripts/reproduce_fig5_from_source_data.py",
    "figures/main/Fig5_reproduced_from_source_data.png",
    "Compact sampled SHAP-interaction points and category interaction edges under data/source_data/main/Fig5/.",
    "Reproducible from sampled plot-ready source data",
    "The full SHAP/validation matrices are not stored in GitHub; the released source data contain sampled plot-ready points sufficient to reproduce the displayed figure."
)


# Extended Data figures
for i, desc in [
    (1, "Robustness of source-specific footprint estimates."),
    (2, "Phase-space support for micro-environmental interpretation."),
    (3, "Full marginal-contribution matrix for feature-domain combinations."),
    (4, "Full 2 脳 7 SHAP interaction matrix for road-emission density 脳 wind speed."),
    (5, "Full 2 脳 7 SHAP interaction matrix for road-emission density 脳 forest density."),
    (6, "Full 2 脳 7 SHAP interaction matrix for road-emission density 脳 average frontal area."),
]:
    ext = "xlsx" if i == 1 else "csv"
    add_row(
        f"EDF{i}", "Extended Data figure", f"Extended Data Fig. {i}", desc,
        f"data/source_data/extended_data/Extended_Data_Fig{i}_source_data.{ext}",
        f"code/scripts/reproduce_extended_data_fig{i}_from_source_data.py",
        f"figures/extended_data/Extended_Data_Fig{i}_reproduced_from_source_data.png",
        "Compact plot-ready source data released in the repository.",
        "Reproducible from compact plot-ready source data",
        "" if i == 1 else "Renumbered into the Extended Data figure set."
    )


# Supplementary figures
for i in range(1, 9):
    csv_path = ROOT / f"data/source_data/supplementary/Supplementary_Fig{i}_source_data.csv"
    xlsx_path = ROOT / f"data/source_data/supplementary/Supplementary_Fig{i}_source_data.xlsx"
    if csv_path.exists():
        source = f"data/source_data/supplementary/Supplementary_Fig{i}_source_data.csv"
    elif xlsx_path.exists():
        source = f"data/source_data/supplementary/Supplementary_Fig{i}_source_data.xlsx"
    else:
        source = f"data/source_data/supplementary/Supplementary_Fig{i}_source_data.csv"

    add_row(
        f"SF{i}", "Supplementary figure", f"Supplementary Fig. {i}",
        f"Supplementary figure {i} source-data and reproduction workflow.",
        source,
        f"code/scripts/reproduce_supplementary_fig{i}_from_source_data.py",
        f"figures/supplementary/Supplementary_Fig{i}_reproduced_from_source_data.png",
        "Compact plot-ready source data released in the repository.",
        "Reproducible from compact plot-ready source data",
        ""
    )


# Supplementary tables
for i, desc in [
    (1, "Distance-band rationale and representative supporting literature."),
    (2, "Model predictor-domain summary and feature-category grouping."),
    (3, "Model-performance and validation-summary table."),
]:
    add_row(
        f"ST{i}", "Supplementary table", f"Supplementary Table {i}",
        desc,
        "Supplementary Information",
        "Not applicable",
        "Supplementary Information",
        "Compiled table in the Supplementary Information.",
        "Documented in Supplementary Information",
        "No separate figure reproduction script is required."
    )


# Supplementary data
add_row(
    "SD1", "Supplementary data", "Supplementary Data 1",
    "Feature list, variable definitions and retained predictor inventory.",
    "data/supplementary_data/Supplementary_Data_1.xlsx",
    "Not applicable",
    "data/supplementary_data/Supplementary_Data_1.xlsx",
    "Released Excel workbook.",
    "Released data workbook",
    ""
)

add_row(
    "SD2", "Supplementary data", "Supplementary Data 2",
    "Reproducibility index linking figures, tables, supplementary data and model-rerun workflows to source data, scripts and expected outputs.",
    "data/supplementary_data/Supplementary_Data_2.xlsx",
    "code/scripts/generate_supplementary_data_2_index_052314.py",
    "data/supplementary_data/Supplementary_Data_2.xlsx",
    "Generated from the current repository structure.",
    "Generated reproducibility index",
    "This workbook is generated by the present script."
)


# Model rerun workflows
for model, script_dir, config_prefix in [
    ("XGBoost", "code/pipelines/xgboost", "xgboost"),
    ("Random Forest", "code/pipelines/rf", "rf"),
    ("LightGBM", "code/pipelines/lgbm", "lgbm"),
]:
    for track in ["spatial", "temporal"]:
        add_row(
            f"MODEL_{model.replace(' ', '_')}_{track}".upper(),
            "Model rerun workflow",
            f"Model rerun: {model} {track} CV",
            f"{model} {track} cross-validation rerun workflow.",
            "data/model_feature_lists/",
            f"{script_dir}/run_{track}_cv.py",
            f"data/model_outputs/{model.lower().replace(' ', '_')}_{track}/",
            f"Model-input matrices are hosted externally when too large for GitHub; configs are expected under code/config/{config_prefix}_{track}_*.json.",
            "Workflow-level reproducibility",
            "Large model-input matrices should be obtained from the archived external release described in Data Availability."
        )


# Submission-facing status corrections for Supplementary Data 2
# These corrections avoid misleading "No" entries for self-generated SD2
# and for model outputs that are intentionally not bundled in GitHub.
for r in rows:
    if r["Item_ID"] == "SD2":
        r["Source_Data_Exists"] = "Yes"
        r["Output_Exists"] = "Yes"
        r["Source_Data_Size_MB"] = ""
        r["Output_Size_MB"] = ""
        r["Notes"] = "Generated by this script; existence status is set after generation."

    if r["Item_Type"] == "Model rerun workflow":
        r["Expected_Output_Path"] = "Not bundled"
        r["Output_Exists"] = "Not bundled"
        r["Output_Size_MB"] = ""
        note = "Model outputs are generated by rerunning the workflow and are not bundled in the release package."
        r["Notes"] = (str(r["Notes"]).strip() + " " + note).strip()


summary_rows = [
    ["Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ["Repository root", "Repository root of the released GitHub package"],
    ["Main figures", 5],
    ["Extended Data figures", 6],
    ["Supplementary figures", 8],
    ["Supplementary tables", 3],
    ["Supplementary data items", 2],
    ["Model rerun workflows", 6],
    ["Total indexed rows", len(rows)],
]


column_defs = [
    ["Item_ID", "Stable internal identifier used in this workbook."],
    ["Item_Type", "Main figure, Extended Data figure, Supplementary figure, Supplementary table, Supplementary data or model rerun workflow."],
    ["Current_Manuscript_Label", "Label used in the current Nature Cities manuscript or Supplementary Information."],
    ["Description", "Brief description of the item."],
    ["Source_Data_Path", "Repository-relative path to the released source data or source-data directory."],
    ["Reproduction_Script_Path", "Repository-relative path to the script used to reproduce the item, where applicable."],
    ["Expected_Output_Path", "Repository-relative path to the expected reproduced output."],
    ["Required_Inputs", "Additional information on required inputs."],
    ["Reproducibility_Level", "Short description of the level of reproducibility provided."],
    ["Notes", "Additional notes, including renumbering or size constraints."],
    ["Source_Data_Exists", "Whether the source-data path exists at workbook-generation time."],
    ["Script_Exists", "Whether the reproduction script exists at workbook-generation time."],
    ["Output_Exists", "Whether the expected output exists at workbook-generation time."],
    ["Source_Data_Size_MB", "Source-data file size in MB if the source path is a file."],
    ["Output_Size_MB", "Output file size in MB if the output path is a file."],
]


wb = Workbook()
ws = wb.active
ws.title = "Reproducibility_Index"

headers = list(rows[0].keys())
ws.append(headers)
for r in rows:
    ws.append([r[h] for h in headers])

ws_summary = wb.create_sheet("Summary")
for row in summary_rows:
    ws_summary.append(row)

ws_defs = wb.create_sheet("Column_Definitions")
ws_defs.append(["Column", "Definition"])
for row in column_defs:
    ws_defs.append(row)


# Styling
header_fill = PatternFill("solid", fgColor="D9EAF7")
bold_font = Font(bold=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for sheet in wb.worksheets:
    for cell in sheet[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, col in enumerate(sheet.columns, start=1):
        max_len = 0
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        width = min(max(max_len + 2, 12), 55)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(OUT_XLSX)
print(f"[OK] Wrote {OUT_XLSX}")
print(f"[INFO] Indexed rows: {len(rows)}")
